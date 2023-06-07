import os
from pprint import pprint

from openpyxl import Workbook
from openpyxl.drawing.image import Image
# from openpyxl.styles.fonts import Font

import ffmpeg
from ffmpeg import *


class CreateExcel:
    def __init__(self, input_path):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Shot'
        self.input_path = input_path
        self.project_path = self.input_path.split('/production/scan/')

    def set_header(self):
        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]

        for i, title in enumerate(header_list):
            self.ws.cell(row=1, column=i + 1, value=title)

    def get_exr_files(self):
        self.exr_files = []

        for path, dirs, files in os.walk(self.input_path):
            if len(files) > 0:
                files.sort(reverse=False)
                self.exr_files.append(os.path.join(path, files[0]))

        return self.exr_files

    def thumbnail_create(self):
        self.thumbnail_path = os.path.join(self.project_path[0], f'tmp/thumb/{self.project_path[1]}')
        if not os.path.exists(self.thumbnail_path):
            os.makedirs(self.thumbnail_path, exist_ok=True)

        for i, exr in enumerate(self.exr_files):
            file_name = os.path.splitext(os.path.basename(exr))[0]
            if not os.path.isfile(f'{self.thumbnail_path}/{file_name}.jpg'):
                ffmpeg.run(output(input(exr), f'{self.thumbnail_path}/{file_name}.jpg'))

    def get_data(self):
        self.data = []

        for exr in self.exr_files:
            temp_list = []
            scan_path = os.path.dirname(exr)
            file_name = os.path.basename(exr).split('.')
            scan_name = file_name[0]
            # pad = '%0' + str(len(file_name[1])) + 'd'
            ext = file_name[2]
            thumbnail_lists = os.listdir(self.thumbnail_path)
            for thumbnail_list in thumbnail_lists:
                if scan_name == os.path.basename(thumbnail_list).split('.')[0]:
                    temp_list.append(thumbnail_list)

            temp_list.append(scan_path)
            temp_list.append(scan_name)
            # temp_list.append(pad)
            temp_list.append(ext)

            self.data.append(temp_list)

        return self.data

    def insert_data(self):
        for i, data in enumerate(self.data):
            self.ws.append({
                'H': data[1],
                'I': data[2],
                # 'K': data[3],
                'L': data[3]
            })

            # insert thumbnail
            image = Image(os.path.join(self.thumbnail_path, data[0]))
            image.width = 250
            image.height = 150
            col_width = image.width * 50 / 350
            row_height = image.height * 250 / 300

            self.ws.add_image(image, anchor='B' + str(i + 2))
            if i == 0:
                self.ws.column_dimensions['B'].width = col_width
            self.ws.row_dimensions[i + 2].height = row_height
            self.ws.cell(row=i + 2, column=2, value=data[0])

    def save_excel_file(self):
        self.excel_path = os.path.join(self.project_path[0], 'production/excel')
        file_name = f'{self.project_path[1]}.xlsx'
        save_dir_path = os.path.join(self.excel_path, file_name)
        self.wb.save(save_dir_path)


if __name__ == "__main__":
    ce = CreateExcel(r"/TD/show/hanjin/production/scan/20221017_plate_scan")
    ce.set_header()
    ce.get_exr_files()
    ce.thumbnail_create()
    ce.get_data()
    ce.insert_data()
    ce.save_excel_file()

